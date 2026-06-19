"""Inventory & Order Management System — FastAPI + PostgreSQL backend."""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Response, Request
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware

import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from sqlalchemy import select, func, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from database import init_db, get_session, session_scope
from models_sql import (
    User, Category, Supplier, Product, Customer, Order, OrderItem,
    InventoryLog, AuditLog,
)
from schemas import (
    UserCreate, UserLogin, UserOut, TokenResponse, RefreshRequest,
    CategoryIn, CategoryOut,
    SupplierIn, SupplierOut,
    ProductIn, ProductUpdate, ProductOut,
    CustomerIn, CustomerOut,
    OrderIn, OrderOut, OrderItemOut,
    StockIn, StockOut, InventoryLogOut,
    AuditLogOut, BulkDelete,
)
from auth import (
    hash_password, verify_password,
    create_access_token, create_refresh_token, decode_token,
    get_current_user, require_roles,
    set_auth_cookies, clear_auth_cookies,
)
from audit_helper import log_audit
from email_service import send_email, low_stock_html
from seed_data import seed_demo_data

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="Inventory & Order Management API", version="1.0.0")
api = APIRouter(prefix="/api")


# ============================================================
# STARTUP
# ============================================================
@app.on_event("startup")
async def startup():
    await init_db()
    async with session_scope() as db:
        # Seed admin/manager/staff users
        for email, name, role, pwd in [
            (os.environ.get("ADMIN_EMAIL", "admin@inventory.com"), "Administrator", "admin", os.environ.get("ADMIN_PASSWORD", "Admin@123")),
            ("manager@inventory.com", "Manager User", "manager", "Manager@123"),
            ("staff@inventory.com", "Staff User", "staff", "Staff@123"),
        ]:
            existing = (await db.execute(select(User).where(User.email == email))).scalar_one_or_none()
            if not existing:
                db.add(User(name=name, email=email, role=role, password_hash=hash_password(pwd)))
        await db.flush()
        try:
            seeded = await seed_demo_data(db)
            if seeded:
                logger.info("Demo data seeded")
        except Exception as e:
            logger.error(f"Seed failed: {e}")


# ============================================================
# AUTH
# ============================================================
@api.post("/auth/register", response_model=TokenResponse)
async def register(payload: UserCreate, response: Response, db: AsyncSession = Depends(get_session)):
    existing = (await db.execute(select(User).where(User.email == payload.email))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    role = payload.role if payload.role in ("admin", "manager", "staff") else "staff"
    user = User(name=payload.name, email=payload.email, role=role, password_hash=hash_password(payload.password))
    db.add(user); await db.flush()
    user_out = UserOut.model_validate(user)
    await log_audit(db, user_out, "register", "user", user.id, f"New user: {user.email}")
    await db.commit()
    access = create_access_token(user.id, user.role)
    refresh = create_refresh_token(user.id)
    set_auth_cookies(response, access, refresh)
    return TokenResponse(access_token=access, refresh_token=refresh, user=user_out)


@api.post("/auth/login", response_model=TokenResponse)
async def login(payload: UserLogin, response: Response, db: AsyncSession = Depends(get_session)):
    user = (await db.execute(select(User).where(User.email == payload.email))).scalar_one_or_none()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    user_out = UserOut.model_validate(user)
    await log_audit(db, user_out, "login", "user", user.id, "User logged in")
    await db.commit()
    access = create_access_token(user.id, user.role)
    refresh = create_refresh_token(user.id)
    set_auth_cookies(response, access, refresh)
    return TokenResponse(access_token=access, refresh_token=refresh, user=user_out)


@api.post("/auth/refresh")
async def refresh(request: Request, response: Response, db: AsyncSession = Depends(get_session)):
    body_token = None
    try:
        body = await request.json()
        body_token = body.get("refresh_token") if isinstance(body, dict) else None
    except Exception:
        body_token = None
    token = body_token or request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token provided")
    data = decode_token(token)
    if data.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid refresh token")
    user = (await db.execute(select(User).where(User.id == data["sub"]))).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    access = create_access_token(user.id, user.role)
    refresh_token = create_refresh_token(user.id)
    set_auth_cookies(response, access, refresh_token)
    return {"access_token": access, "refresh_token": refresh_token}


@api.post("/auth/logout")
async def logout(response: Response, user: UserOut = Depends(get_current_user), db: AsyncSession = Depends(get_session)):
    await log_audit(db, user, "logout", "user", user.id, "User logged out")
    await db.commit()
    clear_auth_cookies(response)
    return {"message": "Logged out"}


@api.get("/auth/profile", response_model=UserOut)
async def profile(user: UserOut = Depends(get_current_user)):
    return user


# ============================================================
# CATEGORIES
# ============================================================
@api.get("/categories", response_model=List[CategoryOut])
async def list_categories(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    rows = (await db.execute(select(Category).order_by(Category.created_at.desc()))).scalars().all()
    return rows


@api.post("/categories", response_model=CategoryOut)
async def create_category(payload: CategoryIn, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    existing = (await db.execute(select(Category).where(Category.name == payload.name))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="Category name already exists")
    cat = Category(**payload.model_dump())
    db.add(cat); await db.flush()
    await log_audit(db, user, "create", "category", cat.id, f"Created category: {cat.name}")
    await db.commit(); await db.refresh(cat)
    return cat


@api.put("/categories/{cat_id}", response_model=CategoryOut)
async def update_category(cat_id: str, payload: CategoryIn, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    cat = (await db.execute(select(Category).where(Category.id == cat_id))).scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    for k, v in payload.model_dump().items():
        setattr(cat, k, v)
    await log_audit(db, user, "update", "category", cat_id, f"Updated category: {cat.name}")
    await db.commit(); await db.refresh(cat)
    return cat


@api.delete("/categories/{cat_id}")
async def delete_category(cat_id: str, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    cat = (await db.execute(select(Category).where(Category.id == cat_id))).scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    name = cat.name
    await db.delete(cat)
    await log_audit(db, user, "delete", "category", cat_id, f"Deleted category: {name}")
    await db.commit()
    return {"message": "Deleted"}


# ============================================================
# SUPPLIERS
# ============================================================
@api.get("/suppliers", response_model=List[SupplierOut])
async def list_suppliers(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    rows = (await db.execute(select(Supplier).order_by(Supplier.created_at.desc()))).scalars().all()
    return rows


@api.get("/suppliers/{sup_id}")
async def supplier_detail(sup_id: str, db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    sup = (await db.execute(select(Supplier).where(Supplier.id == sup_id))).scalar_one_or_none()
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    products = (await db.execute(select(Product).where(Product.supplier_id == sup_id))).scalars().all()
    logs = (await db.execute(
        select(InventoryLog).where(InventoryLog.supplier_id == sup_id).order_by(InventoryLog.created_at.desc()).limit(200)
    )).scalars().all()
    return {
        "supplier": SupplierOut.model_validate(sup),
        "products": [ProductOut.model_validate(p) for p in products],
        "transactions": [InventoryLogOut.model_validate(l) for l in logs],
    }


@api.post("/suppliers", response_model=SupplierOut)
async def create_supplier(payload: SupplierIn, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    sup = Supplier(**payload.model_dump())
    db.add(sup); await db.flush()
    await log_audit(db, user, "create", "supplier", sup.id, f"Created supplier: {sup.name}")
    await db.commit(); await db.refresh(sup)
    return sup


@api.put("/suppliers/{sup_id}", response_model=SupplierOut)
async def update_supplier(sup_id: str, payload: SupplierIn, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    sup = (await db.execute(select(Supplier).where(Supplier.id == sup_id))).scalar_one_or_none()
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    for k, v in payload.model_dump().items():
        setattr(sup, k, v)
    await log_audit(db, user, "update", "supplier", sup_id, f"Updated supplier: {sup.name}")
    await db.commit(); await db.refresh(sup)
    return sup


@api.delete("/suppliers/{sup_id}")
async def delete_supplier(sup_id: str, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    sup = (await db.execute(select(Supplier).where(Supplier.id == sup_id))).scalar_one_or_none()
    if not sup:
        raise HTTPException(status_code=404, detail="Supplier not found")
    name = sup.name
    await db.delete(sup)
    await log_audit(db, user, "delete", "supplier", sup_id, f"Deleted supplier: {name}")
    await db.commit()
    return {"message": "Deleted"}


# ============================================================
# PRODUCTS
# ============================================================
@api.get("/products")
async def list_products(
    search: Optional[str] = None,
    category_id: Optional[str] = None,
    supplier_id: Optional[str] = None,
    stock_status: Optional[str] = None,
    sort: str = "created_at",
    order: str = "desc",
    page: int = 1,
    limit: int = 20,
    db: AsyncSession = Depends(get_session),
    _: UserOut = Depends(get_current_user),
):
    q = select(Product)
    if search:
        like = f"%{search}%"
        q = q.where(or_(Product.name.ilike(like), Product.sku.ilike(like), Product.barcode.ilike(like)))
    if category_id:
        q = q.where(Product.category_id == category_id)
    if supplier_id:
        q = q.where(Product.supplier_id == supplier_id)
    if stock_status == "low":
        q = q.where(and_(Product.quantity > 0, Product.quantity <= Product.reorder_level))
    elif stock_status == "out":
        q = q.where(Product.quantity == 0)
    elif stock_status == "in_stock":
        q = q.where(Product.quantity > Product.reorder_level)

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar() or 0
    sort_col = getattr(Product, sort, Product.created_at)
    sort_col = sort_col.desc() if order == "desc" else sort_col.asc()
    q = q.order_by(sort_col).offset((page - 1) * limit).limit(limit)
    items = (await db.execute(q)).scalars().all()
    return {
        "items": [ProductOut.model_validate(p).model_dump(mode="json") for p in items],
        "total": total, "page": page, "limit": limit,
    }


@api.get("/products/{pid}", response_model=ProductOut)
async def get_product(pid: str, db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    p = (await db.execute(select(Product).where(Product.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    return p


@api.post("/products", response_model=ProductOut, status_code=201)
async def create_product(payload: ProductIn, db: AsyncSession = Depends(get_session),
                         user: UserOut = Depends(require_roles("admin", "manager"))):
    existing = (await db.execute(select(Product).where(Product.sku == payload.sku))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="SKU already exists")
    data = payload.model_dump()
    # Normalize "" → None for FK fields
    if not data.get("category_id"): data["category_id"] = None
    if not data.get("supplier_id"): data["supplier_id"] = None
    p = Product(**data)
    db.add(p); await db.flush()
    await log_audit(db, user, "create", "product", p.id, f"Created product: {p.name} ({p.sku})")
    await db.commit(); await db.refresh(p)
    return p


@api.put("/products/{pid}", response_model=ProductOut)
async def update_product(pid: str, payload: ProductUpdate, db: AsyncSession = Depends(get_session),
                         user: UserOut = Depends(require_roles("admin", "manager"))):
    p = (await db.execute(select(Product).where(Product.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    changes = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "sku" in changes and changes["sku"] != p.sku:
        existing = (await db.execute(select(Product).where(Product.sku == changes["sku"]))).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=400, detail="SKU already exists")
    if "quantity" in changes and changes["quantity"] < 0:
        raise HTTPException(status_code=400, detail="Product quantity cannot be negative")
    for k, v in changes.items():
        setattr(p, k, v)
    await log_audit(db, user, "update", "product", pid, f"Updated product: {p.name}")
    await db.commit(); await db.refresh(p)
    return p


@api.delete("/products/{pid}")
async def delete_product(pid: str, db: AsyncSession = Depends(get_session),
                         user: UserOut = Depends(require_roles("admin", "manager"))):
    p = (await db.execute(select(Product).where(Product.id == pid))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    name = p.name
    await db.delete(p)
    await log_audit(db, user, "delete", "product", pid, f"Deleted product: {name}")
    await db.commit()
    return {"message": "Deleted"}


@api.post("/products/bulk-delete")
async def bulk_delete_products(payload: BulkDelete, db: AsyncSession = Depends(get_session),
                               user: UserOut = Depends(require_roles("admin", "manager"))):
    rows = (await db.execute(select(Product).where(Product.id.in_(payload.ids)))).scalars().all()
    deleted = 0
    for r in rows:
        await db.delete(r); deleted += 1
    await log_audit(db, user, "bulk_delete", "product", None, f"Bulk deleted {deleted} products")
    await db.commit()
    return {"deleted": deleted}


# ============================================================
# CUSTOMERS
# ============================================================
@api.get("/customers", response_model=List[CustomerOut])
async def list_customers(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    rows = (await db.execute(select(Customer).order_by(Customer.created_at.desc()))).scalars().all()
    return rows


@api.get("/customers/{cid}", response_model=CustomerOut)
async def get_customer(cid: str, db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    c = (await db.execute(select(Customer).where(Customer.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    return c


@api.post("/customers", response_model=CustomerOut, status_code=201)
async def create_customer(payload: CustomerIn, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(get_current_user)):
    existing = (await db.execute(select(Customer).where(Customer.email == payload.email))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="Customer email already exists")
    c = Customer(**payload.model_dump())
    db.add(c); await db.flush()
    await log_audit(db, user, "create", "customer", c.id, f"Created customer: {c.full_name}")
    await db.commit(); await db.refresh(c)
    return c


@api.put("/customers/{cid}", response_model=CustomerOut)
async def update_customer(cid: str, payload: CustomerIn, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(get_current_user)):
    c = (await db.execute(select(Customer).where(Customer.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    if payload.email != c.email:
        existing = (await db.execute(select(Customer).where(Customer.email == payload.email))).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=400, detail="Customer email already exists")
    for k, v in payload.model_dump().items():
        setattr(c, k, v)
    await log_audit(db, user, "update", "customer", cid, f"Updated customer: {c.full_name}")
    await db.commit(); await db.refresh(c)
    return c


@api.delete("/customers/{cid}")
async def delete_customer(cid: str, db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    c = (await db.execute(select(Customer).where(Customer.id == cid))).scalar_one_or_none()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    name = c.full_name
    await db.delete(c)
    await log_audit(db, user, "delete", "customer", cid, f"Deleted customer: {name}")
    await db.commit()
    return {"message": "Deleted"}


# ============================================================
# ORDERS
# ============================================================
async def _generate_order_number(db: AsyncSession) -> str:
    count = (await db.execute(select(func.count()).select_from(Order))).scalar() or 0
    return f"ORD-{1000 + count + 1:05d}"


@api.get("/orders", response_model=List[OrderOut])
async def list_orders(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    rows = (await db.execute(
        select(Order).options(selectinload(Order.items), selectinload(Order.customer)).order_by(Order.created_at.desc())
    )).scalars().all()
    return rows


@api.get("/orders/{oid}", response_model=OrderOut)
async def get_order(oid: str, db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    o = (await db.execute(
        select(Order).options(selectinload(Order.items), selectinload(Order.customer)).where(Order.id == oid)
    )).scalar_one_or_none()
    if not o:
        raise HTTPException(status_code=404, detail="Order not found")
    return o


@api.post("/orders", response_model=OrderOut, status_code=201)
async def create_order(payload: OrderIn, db: AsyncSession = Depends(get_session),
                       user: UserOut = Depends(get_current_user)):
    customer = (await db.execute(select(Customer).where(Customer.id == payload.customer_id))).scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Validate + aggregate (handle duplicate product ids by merging quantities)
    product_qty: dict[str, int] = {}
    for item in payload.items:
        product_qty[item.product_id] = product_qty.get(item.product_id, 0) + item.quantity

    products = (await db.execute(select(Product).where(Product.id.in_(product_qty.keys())))).scalars().all()
    product_map = {p.id: p for p in products}

    # Validate existence + stock
    items_out: list[OrderItem] = []
    total = 0.0
    for pid, qty in product_qty.items():
        p = product_map.get(pid)
        if not p:
            raise HTTPException(status_code=404, detail=f"Product not found: {pid}")
        if p.quantity < qty:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock for {p.name} ({p.sku}). Available: {p.quantity}, requested: {qty}",
            )
        line = round(p.selling_price * qty, 2)
        total += line
        items_out.append(OrderItem(
            product_id=p.id, product_name=p.name, product_sku=p.sku,
            unit_price=p.selling_price, quantity=qty, line_total=line,
        ))

    # All validations passed — decrement stock + log inventory movement
    for pid, qty in product_qty.items():
        p = product_map[pid]
        prev = p.quantity
        p.quantity = prev - qty
        db.add(InventoryLog(
            product_id=p.id, product_name=p.name,
            action="stock_out", quantity_change=-qty,
            previous_quantity=prev, new_quantity=p.quantity,
            user_id=user.id, user_name=user.name,
            destination=f"Order for {customer.full_name}",
            notes=f"Auto-deducted on order creation",
        ))

    order = Order(
        order_number=await _generate_order_number(db),
        customer_id=customer.id,
        total_amount=round(total, 2),
        status="pending", notes=payload.notes or "",
        items=items_out,
    )
    db.add(order); await db.flush()
    await log_audit(db, user, "create", "order", order.id,
                    f"Created order {order.order_number} ({len(items_out)} items, total ${total:.2f})")
    await db.commit()

    # Re-fetch with relationships
    order = (await db.execute(
        select(Order).options(selectinload(Order.items), selectinload(Order.customer)).where(Order.id == order.id)
    )).scalar_one()
    return order


@api.delete("/orders/{oid}")
async def cancel_order(oid: str, db: AsyncSession = Depends(get_session),
                       user: UserOut = Depends(get_current_user)):
    order = (await db.execute(
        select(Order).options(selectinload(Order.items)).where(Order.id == oid)
    )).scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status == "cancelled":
        raise HTTPException(status_code=400, detail="Order already cancelled")

    # Restock items
    for it in order.items:
        p = (await db.execute(select(Product).where(Product.id == it.product_id))).scalar_one_or_none()
        if not p:
            continue
        prev = p.quantity
        p.quantity = prev + it.quantity
        db.add(InventoryLog(
            product_id=p.id, product_name=p.name,
            action="stock_in", quantity_change=it.quantity,
            previous_quantity=prev, new_quantity=p.quantity,
            user_id=user.id, user_name=user.name,
            notes=f"Order {order.order_number} cancelled — restocked",
        ))
    order.status = "cancelled"
    await log_audit(db, user, "cancel", "order", oid, f"Cancelled order {order.order_number}")
    await db.commit()
    return {"message": "Order cancelled"}


# ============================================================
# INVENTORY
# ============================================================
@api.post("/inventory/stock-in", response_model=InventoryLogOut)
async def stock_in(payload: StockIn, db: AsyncSession = Depends(get_session),
                   user: UserOut = Depends(require_roles("admin", "manager", "staff"))):
    p = (await db.execute(select(Product).where(Product.id == payload.product_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    prev = p.quantity
    p.quantity = prev + payload.quantity
    log = InventoryLog(
        product_id=p.id, product_name=p.name,
        action="stock_in", quantity_change=payload.quantity,
        previous_quantity=prev, new_quantity=p.quantity,
        user_id=user.id, user_name=user.name,
        supplier_id=payload.supplier_id, purchase_cost=payload.purchase_cost,
        notes=payload.notes or "",
    )
    db.add(log); await db.flush()
    await log_audit(db, user, "stock_in", "inventory", p.id, f"+{payload.quantity} {p.name}")
    await db.commit(); await db.refresh(log)
    return log


@api.post("/inventory/stock-out", response_model=InventoryLogOut)
async def stock_out(payload: StockOut, db: AsyncSession = Depends(get_session),
                    user: UserOut = Depends(require_roles("admin", "manager", "staff"))):
    p = (await db.execute(select(Product).where(Product.id == payload.product_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    if p.quantity < payload.quantity:
        raise HTTPException(status_code=400, detail="Insufficient stock")
    prev = p.quantity
    p.quantity = prev - payload.quantity
    log = InventoryLog(
        product_id=p.id, product_name=p.name,
        action="stock_out", quantity_change=-payload.quantity,
        previous_quantity=prev, new_quantity=p.quantity,
        user_id=user.id, user_name=user.name,
        destination=payload.destination or "", notes=payload.notes or "",
    )
    db.add(log); await db.flush()
    await log_audit(db, user, "stock_out", "inventory", p.id, f"-{payload.quantity} {p.name}")
    await db.commit(); await db.refresh(log)
    try:
        if p.quantity <= p.reorder_level:
            recipient = os.environ.get("ALERT_RECIPIENT")
            if recipient:
                await send_email(recipient, f"Low Stock Alert: {p.name}",
                                 low_stock_html([{"name": p.name, "sku": p.sku, "quantity": p.quantity, "reorder_level": p.reorder_level}]))
    except Exception as e:
        logger.error(f"Low stock alert failed: {e}")
    return log


@api.get("/inventory/history", response_model=List[InventoryLogOut])
async def inventory_history(product_id: Optional[str] = None, action: Optional[str] = None,
                            limit: int = 100, db: AsyncSession = Depends(get_session),
                            _: UserOut = Depends(get_current_user)):
    q = select(InventoryLog)
    if product_id: q = q.where(InventoryLog.product_id == product_id)
    if action: q = q.where(InventoryLog.action == action)
    rows = (await db.execute(q.order_by(InventoryLog.created_at.desc()).limit(limit))).scalars().all()
    return rows


# ============================================================
# REPORTS / DASHBOARD
# ============================================================
async def _compute_kpis(db: AsyncSession) -> dict:
    total_products = (await db.execute(select(func.count()).select_from(Product))).scalar() or 0
    total_categories = (await db.execute(select(func.count()).select_from(Category))).scalar() or 0
    total_suppliers = (await db.execute(select(func.count()).select_from(Supplier))).scalar() or 0
    total_customers = (await db.execute(select(func.count()).select_from(Customer))).scalar() or 0
    total_orders = (await db.execute(select(func.count()).select_from(Order))).scalar() or 0
    low_stock = (await db.execute(select(func.count()).select_from(Product).where(
        and_(Product.quantity > 0, Product.quantity <= Product.reorder_level)
    ))).scalar() or 0
    out_of_stock = (await db.execute(select(func.count()).select_from(Product).where(Product.quantity == 0))).scalar() or 0
    inv_value_row = (await db.execute(select(func.coalesce(func.sum(Product.cost_price * Product.quantity), 0.0)))).scalar() or 0.0

    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    todays_tx = (await db.execute(select(func.count()).select_from(InventoryLog).where(InventoryLog.created_at >= today))).scalar() or 0

    return {
        "total_products": total_products,
        "total_categories": total_categories,
        "total_suppliers": total_suppliers,
        "total_customers": total_customers,
        "total_orders": total_orders,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "inventory_value": round(float(inv_value_row), 2),
        "todays_transactions": todays_tx,
    }


async def _stock_trend(db: AsyncSession, days: int = 14) -> list:
    out = []
    today_utc = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(days - 1, -1, -1):
        start = today_utc - timedelta(days=i)
        end = start + timedelta(days=1)
        ins = (await db.execute(select(func.count()).select_from(InventoryLog).where(
            and_(InventoryLog.action == "stock_in", InventoryLog.created_at >= start, InventoryLog.created_at < end)
        ))).scalar() or 0
        outs = (await db.execute(select(func.count()).select_from(InventoryLog).where(
            and_(InventoryLog.action == "stock_out", InventoryLog.created_at >= start, InventoryLog.created_at < end)
        ))).scalar() or 0
        out.append({"date": start.strftime("%m-%d"), "stock_in": ins, "stock_out": outs})
    return out


async def _category_distribution(db: AsyncSession) -> list:
    rows = (await db.execute(
        select(Category.name, func.count(Product.id))
        .select_from(Product).join(Category, Product.category_id == Category.id, isouter=True)
        .group_by(Category.name)
    )).all()
    return [{"name": name or "Uncategorized", "value": count} for name, count in rows]


async def _supplier_contribution(db: AsyncSession) -> list:
    rows = (await db.execute(
        select(Supplier.name, func.count(Product.id))
        .select_from(Product).join(Supplier, Product.supplier_id == Supplier.id)
        .group_by(Supplier.name)
    )).all()
    return [{"name": name, "value": count} for name, count in rows]


async def _monthly_growth(db: AsyncSession, months: int = 6) -> list:
    out = []
    now = datetime.now(timezone.utc)
    for i in range(months - 1, -1, -1):
        m_start = (now.replace(day=1) - timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # rough next-month boundary
        next_m = (m_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        count = (await db.execute(select(func.count()).select_from(Product).where(
            and_(Product.created_at >= m_start, Product.created_at < next_m)
        ))).scalar() or 0
        out.append({"month": m_start.strftime("%b %Y"), "products": count})
    return out


@api.get("/reports/dashboard")
async def dashboard(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    return {
        "kpi": await _compute_kpis(db),
        "stock_trend": await _stock_trend(db),
        "category_distribution": await _category_distribution(db),
        "monthly_growth": await _monthly_growth(db),
        "supplier_contribution": await _supplier_contribution(db),
    }


@api.get("/reports/low-stock")
async def low_stock_report(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    low = (await db.execute(select(Product).where(and_(Product.quantity > 0, Product.quantity <= Product.reorder_level)))).scalars().all()
    out = (await db.execute(select(Product).where(Product.quantity == 0))).scalars().all()
    return {
        "low_stock": [ProductOut.model_validate(p).model_dump(mode="json") for p in low],
        "out_of_stock": [ProductOut.model_validate(p).model_dump(mode="json") for p in out],
    }


@api.post("/reports/low-stock/send-alert")
async def send_low_stock_alert(db: AsyncSession = Depends(get_session),
                               user: UserOut = Depends(require_roles("admin", "manager"))):
    products = (await db.execute(select(Product).where(Product.quantity <= Product.reorder_level))).scalars().all()
    if not products:
        return {"sent": False, "message": "No low-stock products"}
    recipient = os.environ.get("ALERT_RECIPIENT")
    if not recipient:
        raise HTTPException(status_code=400, detail="No alert recipient configured")
    items = [{"name": p.name, "sku": p.sku, "quantity": p.quantity, "reorder_level": p.reorder_level} for p in products]
    ok = await send_email(recipient, "Low Stock Alert - Inventory Ops", low_stock_html(items))
    await log_audit(db, user, "send_alert", "report", None, f"Low-stock alert sent ({len(products)} items)")
    await db.commit()
    return {"sent": ok, "count": len(products), "recipient": recipient}


@api.get("/reports/transactions")
async def tx_report(range: str = "weekly", db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    if range == "daily": start = now - timedelta(days=1)
    elif range == "monthly": start = now - timedelta(days=30)
    else: start = now - timedelta(days=7)
    items = (await db.execute(
        select(InventoryLog).where(InventoryLog.created_at >= start).order_by(InventoryLog.created_at.desc()).limit(5000)
    )).scalars().all()
    return {"items": [InventoryLogOut.model_validate(i).model_dump(mode="json") for i in items], "count": len(items)}


@api.get("/reports/monthly")
async def monthly_report(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    items = (await db.execute(
        select(InventoryLog).where(InventoryLog.created_at >= start).order_by(InventoryLog.created_at.desc()).limit(5000)
    )).scalars().all()
    in_total = sum(i.quantity_change for i in items if i.action == "stock_in")
    out_total = sum(-i.quantity_change for i in items if i.action == "stock_out")
    return {
        "month": now.strftime("%B %Y"),
        "total_transactions": len(items),
        "stock_in_units": in_total,
        "stock_out_units": out_total,
        "items": [InventoryLogOut.model_validate(i).model_dump(mode="json") for i in items],
    }


# ============================================================
# EXPORTS
# ============================================================
@api.get("/products/export/excel")
async def export_products_excel(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    items = (await db.execute(select(Product).order_by(Product.created_at.desc()))).scalars().all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Products"
    ws.append(["SKU", "Name", "Description", "Quantity", "Cost Price", "Selling Price", "Reorder Level", "Barcode"])
    for p in items:
        ws.append([p.sku, p.name, p.description, p.quantity, p.cost_price, p.selling_price, p.reorder_level, p.barcode])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=products.xlsx"})


@api.get("/products/export/pdf")
async def export_products_pdf(db: AsyncSession = Depends(get_session), _: UserOut = Depends(get_current_user)):
    items = (await db.execute(select(Product).order_by(Product.created_at.desc()))).scalars().all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Products Report", styles["Title"]), Spacer(1, 12)]
    data = [["SKU", "Name", "Qty", "Cost", "Price", "Reorder"]]
    for p in items:
        data.append([p.sku, p.name, str(p.quantity), f"{p.cost_price:.2f}", f"{p.selling_price:.2f}", str(p.reorder_level)])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10B981")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(t); doc.build(elements); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=products.pdf"})


@api.post("/products/import")
async def import_products(file: UploadFile = File(...), db: AsyncSession = Depends(get_session),
                          user: UserOut = Depends(require_roles("admin", "manager"))):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty sheet")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    if not {"sku", "name"}.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="Required columns: sku, name")

    created = skipped = 0
    for row in rows[1:]:
        if not any(row): continue
        data = dict(zip(headers, row))
        sku = str(data.get("sku") or "").strip()
        name = str(data.get("name") or "").strip()
        if not sku or not name: skipped += 1; continue
        existing = (await db.execute(select(Product).where(Product.sku == sku))).scalar_one_or_none()
        if existing: skipped += 1; continue
        db.add(Product(
            sku=sku, name=name,
            description=str(data.get("description") or ""),
            cost_price=float(data.get("cost_price") or 0),
            selling_price=float(data.get("selling_price") or 0),
            quantity=int(data.get("quantity") or 0),
            reorder_level=int(data.get("reorder_level") or 10),
            barcode=str(data.get("barcode") or ""),
        ))
        created += 1
    await log_audit(db, user, "import", "product", None, f"Imported {created}, skipped {skipped}")
    await db.commit()
    return {"created": created, "skipped": skipped}


@api.get("/reports/export/excel")
async def reports_export_excel(report: str = "products", db: AsyncSession = Depends(get_session),
                               _: UserOut = Depends(get_current_user)):
    wb = openpyxl.Workbook(); ws = wb.active
    if report == "low_stock":
        ws.title = "Low Stock"
        ws.append(["SKU", "Name", "Quantity", "Reorder Level"])
        for p in (await db.execute(select(Product).where(Product.quantity <= Product.reorder_level))).scalars().all():
            ws.append([p.sku, p.name, p.quantity, p.reorder_level])
    elif report == "transactions":
        ws.title = "Transactions"
        ws.append(["Date", "Action", "Product", "Qty Change", "Previous", "New", "User", "Notes"])
        for row in (await db.execute(select(InventoryLog).order_by(InventoryLog.created_at.desc()).limit(5000))).scalars().all():
            ws.append([row.created_at.isoformat(), row.action, row.product_name, row.quantity_change,
                       row.previous_quantity, row.new_quantity, row.user_name, row.notes or ""])
    elif report == "audit":
        ws.title = "Audit"
        ws.append(["Date", "User", "Action", "Entity", "Details"])
        for row in (await db.execute(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(5000))).scalars().all():
            ws.append([row.created_at.isoformat(), row.user_name, row.action, row.entity, row.details or ""])
    else:
        ws.title = "Products"
        ws.append(["SKU", "Name", "Quantity", "Cost", "Price", "Reorder"])
        for p in (await db.execute(select(Product))).scalars().all():
            ws.append([p.sku, p.name, p.quantity, p.cost_price, p.selling_price, p.reorder_level])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={report}.xlsx"})


# ============================================================
# AUDIT
# ============================================================
@api.get("/audit-logs", response_model=List[AuditLogOut])
async def list_audit(entity: Optional[str] = None, action: Optional[str] = None, limit: int = 200,
                     db: AsyncSession = Depends(get_session),
                     _: UserOut = Depends(require_roles("admin", "manager"))):
    q = select(AuditLog)
    if entity: q = q.where(AuditLog.entity == entity)
    if action: q = q.where(AuditLog.action == action)
    rows = (await db.execute(q.order_by(AuditLog.created_at.desc()).limit(limit))).scalars().all()
    return rows


# ============================================================
# MOUNT
# ============================================================
@api.get("/")
async def root():
    return {"message": "Inventory & Order Management API", "status": "online", "version": "1.0.0"}


@api.get("/healthz")
async def healthz():
    return {"status": "ok"}


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origin_regex=os.environ.get("CORS_ORIGIN_REGEX", ".*"),
    allow_methods=["*"],
    allow_headers=["*"],
)
